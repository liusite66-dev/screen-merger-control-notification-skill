#!/usr/bin/env python3
"""Deterministic preparation and reporting for China merger-control screening."""
from __future__ import annotations
import argparse, hashlib, json, os, re, shutil, sys, zipfile, html, subprocess, tempfile
from collections import Counter
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

MARKER='.merger-control-workspace'; LEVELS={'未发现明显问题':0,'低风险':1,'待核验':1,'中风险':2,'高风险':3}
THRESHOLDS={'global_total':120,'china_two':40,'china_one':15} # 亿元; configurable and visibly labelled
class PipelineError(RuntimeError): pass
def sha(p):
 h=hashlib.sha256();
 with Path(p).open('rb') as f:
  for b in iter(lambda:f.read(1024*1024),b''): h.update(b)
 return h.hexdigest()
def workspace(p):
 p=Path(p).expanduser().resolve()
 if p.exists() and any(p.iterdir()): raise PipelineError('工作目录必须为空')
 p.mkdir(parents=True,mode=0o700,exist_ok=True); (p/MARKER).write_text('temporary\n'); (p/'markdown').mkdir(mode=0o700); return p
def cleanup(p):
 p=Path(p).resolve()
 if not (p/MARKER).is_file() or len(p.parts)<4: raise PipelineError('拒绝清理非受管目录')
 shutil.rmtree(p)
def read_json(arg): return json.load(sys.stdin) if arg=='-' else json.loads(Path(arg).read_text(encoding='utf-8'))
def read_xlsx(path):
 book=load_workbook(path,read_only=True,data_only=True); out=[]
 try:
  for ws in book.worksheets:
   rows=list(ws.iter_rows(values_only=True));
   if not rows: continue
   heads=[str(x or '').strip() for x in rows[0]]
   for ri,row in enumerate(rows[1:],2):
    d={heads[i]: row[i] for i in range(min(len(heads),len(row))) if heads[i]}
    if any(v not in (None,'') for v in d.values()): d['_source']=f'{Path(path).name}/{ws.title}/第{ri}行'; out.append(d)
 finally: book.close()
 return out
def material(p,ws,i):
 p=Path(p).resolve(); data={'material_id':f'M-{i:03d}','source_name':p.name,'source_path':str(p),'sha256':sha(p),'format':p.suffix.lower(),'status':'已记录'}
 if p.suffix.lower() in {'.txt','.md','.html','.htm'}: text=p.read_text(encoding='utf-8',errors='replace'); note='文本读取'
 elif p.suffix.lower()=='.docx':
  try:
   with zipfile.ZipFile(p) as z: text=' '.join(ET.fromstring(z.read('word/document.xml')).itertext()); note='DOCX OOXML文本提取'
  except Exception: text=''; note='DOCX解析失败'
 elif p.suffix.lower()=='.pdf':
  text=''; note='PDF待MarkItDown转换'
 else: text=''; note='待MarkItDown转换/OCR'
 if not text:
  try:
   try:
    from markitdown import MarkItDown
   except ImportError:
    dep=ws/'markitdown-deps'; dep.mkdir(mode=0o700,exist_ok=True)
    subprocess.run([sys.executable,'-m','pip','install','--disable-pip-version-check','--quiet','--target',str(dep),'markitdown'],check=True,capture_output=True,timeout=300)
    sys.path.insert(0,str(dep)); from markitdown import MarkItDown
   r=MarkItDown().convert_local(str(p)); text=getattr(r,'markdown',None) or getattr(r,'text_content','') or ''; note='MarkItDown本地转换'
  except Exception as e: note=f'转换失败待核验({type(e).__name__})'
 if text:
  md=ws/'markdown'/f'M-{i:03d}.md'; md.write_text(f'# {p.name}\n\n{text}',encoding='utf-8'); os.chmod(md,0o600); data.update(markdown_path=str(md),char_count=len(text))
 else: data.update(markdown_path='',char_count=0,status='待核验')
 data['processing_note']=note; return data
def prepare(a):
 if a.processing_environment=='cloud' and not a.privacy_confirmed: raise PipelineError('云端读取必须提供--privacy-confirmed')
 tx=Path(a.transaction).resolve();
 if not tx.exists(): raise PipelineError('交易结构文件不存在')
 ws=workspace(a.workspace)
 try:
  raw=read_json(str(tx)) if tx.suffix.lower()=='.json' else {'transaction':read_xlsx(tx)}
  if not isinstance(raw,dict): raise PipelineError('交易结构必须是对象')
  attachments=[]
  for i,p in enumerate(a.attachment or [],1):
   pp=Path(p).resolve();
   if not pp.is_file(): raise PipelineError(f'附件不存在：{pp}')
   attachments.append(material(pp,ws,i))
  bundle={'schema_version':'1.0','created_at':datetime.now().astimezone().isoformat(timespec='seconds'),'transaction_file':str(tx),'transaction_sha256':sha(tx),'transaction':raw,'materials':attachments,'thresholds':THRESHOLDS,'policy':'金额单位亿元；缺失不等于未达到；仅内部初查'}
  out=ws/'merger-bundle.json'; out.write_text(json.dumps(bundle,ensure_ascii=False,indent=2),encoding='utf-8'); os.chmod(out,0o600); print(out)
 except Exception:
  cleanup(ws); raise
def arr(d,k):
 v=d.get(k,[])
 if not isinstance(v,list) or any(not isinstance(x,dict) for x in v): raise PipelineError(f'{k}必须是对象数组')
 return v
def validate(d,b):
 if not isinstance(d,dict): raise PipelineError('analysis-json必须是对象')
 keys=['operators','control_findings','turnover_records','legal_bases','procedure_assessment','material_checklist','review_items','processing_records']
 out={k:arr(d,k) for k in keys}
 for x in out['review_items']:
  if x.get('risk_level') not in LEVELS: raise PipelineError('风险等级无效')
  if x.get('risk_level') in ('高风险','中风险') and not x.get('recommendation'): raise PipelineError('高/中风险必须有recommendation')
 return out
def style(ws):
 for c in ws[1]: c.fill=PatternFill('solid',fgColor='1F4E78'); c.font=Font(color='FFFFFF',bold=True); c.alignment=Alignment(wrap_text=True)
 for row in ws.iter_rows(min_row=2):
  for c in row: c.alignment=Alignment(wrap_text=True,vertical='top')
 ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
 for i in range(1,ws.max_column+1): ws.column_dimensions[get_column_letter(i)].width=min(46,max(12,max(len(str(ws.cell(r,i).value or '')) for r in range(1,min(ws.max_row,80)+1))+2))
def add(book,name,cols,rows):
 ws=book.create_sheet(name); ws.append(cols)
 for x in rows: ws.append([json.dumps(x.get(c),ensure_ascii=False) if isinstance(x.get(c),(dict,list)) else x.get(c,'') for c in cols])
 style(ws)
def summarize(b,d):
 recs=b.get('transaction',{}); tx=recs.get('transaction',recs)
 if isinstance(tx,list): tx={}
 turnovers=[x for x in d['turnover_records'] if isinstance(x.get('amount'),(int,float))]
 total=sum(x['amount'] for x in turnovers if x.get('scope')=='global_total')
 china=sum(x['amount'] for x in turnovers if x.get('scope')=='china_total')
 entities=len({x.get('operator_id') for x in turnovers if x.get('operator_id')})
 threshold='待核验'
 if turnovers: threshold='达到门槛候选' if total>=THRESHOLDS['global_total'] and china>=THRESHOLDS['china_two'] else '未达到法定门槛（按提供数据初筛）'
 if not turnovers: threshold='待核验：缺少可核验营业额'
 closing=str(tx.get('closing_status','') or tx.get('交易阶段',''))
 gun='高风险：已交割/已实施但未确认获批' if any(k in closing.lower() for k in ('closed','已交割','实施')) else '待核验：需确认签约、交割及控制权实施时间'
 return {'threshold_assessment':threshold,'global_total':total,'china_total':china,'operator_count':entities,'gun_jumping':gun}
def report(a):
 bp=Path(a.bundle).resolve(); ws=bp.parent
 if not (ws/MARKER).is_file(): raise PipelineError('bundle不在受管目录')
 b=json.loads(bp.read_text(encoding='utf-8')); d=validate(read_json(a.analysis_json),b); outdir=Path(a.output_dir).resolve(); outdir.mkdir(parents=True,exist_ok=True)
 title=str(b.get('transaction',{}).get('title') or b.get('transaction',{}).get('交易名称') or Path(b['transaction_file']).stem); title=re.sub(r'[\\/:*?"<>|]','_',title)
 xlsx=outdir/f'{title}_经营者集中申报审查.xlsx'; htm=outdir/f'{title}_经营者集中申报审查.html'
 if xlsx.exists() or htm.exists(): raise PipelineError('输出已存在，拒绝覆盖')
 s=summarize(b,d); counts=Counter(x.get('risk_level') for x in d['review_items']); book=Workbook(); ws0=book.active; ws0.title='审查摘要'; ws0.append(['项目','内容'])
 for k,v in [('交易名称',title),('门槛初筛',s['threshold_assessment']),('全球营业额合计(亿元)',s['global_total']),('中国境内营业额合计(亿元)',s['china_total']),('经营者数量',s['operator_count']),('抢跑风险',s['gun_jumping']),*[(k,counts[k]) for k in LEVELS]]: ws0.append([k,v])
 style(ws0)
 cols={'经营者与交易':['operator_id','name','role','transaction_type','source_locator'],'控制权分析':['finding_id','operator_id','control_type','basis','risk_level','source_locator','confidence'],'营业额记录':['operator_id','fiscal_year','scope','amount','currency','source_locator','confidence'],'程序分流':['item_id','route','status','reason','source_locator'],'申报材料清单':['item_id','material','status','owner','source_locator'],'风险与建议':['item_id','topic','risk_level','finding','recommendation','source_locator','confidence'],'法律依据':['basis_id','name','article','version','url','retrieved_at','status'],'待核验':['item_id','question','impact','suggested_owner','source_locator'],'处理记录':['subject','status','method','note']}
 mapping={'经营者与交易':'operators','控制权分析':'control_findings','营业额记录':'turnover_records','程序分流':'procedure_assessment','申报材料清单':'material_checklist','风险与建议':'review_items','法律依据':'legal_bases','待核验':'review_items','处理记录':'processing_records'}
 for n,k in mapping.items(): add(book,n,cols[n],d[k])
 add(book,'材料与哈希',['material_id','source_name','sha256','format','markdown_path','processing_note','status'],b['materials'])
 book.save(xlsx); book.close()
 rows=''.join(f'<tr><td>{html.escape(str(x.get("topic","")))}</td><td>{html.escape(str(x.get("risk_level","")))}</td><td>{html.escape(str(x.get("finding","")))}</td><td>{html.escape(str(x.get("recommendation","")))}</td></tr>' for x in d['review_items'])
 htm.write_text(f'<!doctype html><meta charset="utf-8"><title>{html.escape(title)}</title><style>body{{font-family:Arial,"Microsoft YaHei";margin:30px;color:#18212b}}.cards{{display:flex;gap:12px;flex-wrap:wrap}}.card{{padding:14px;background:#eef4f7;border:1px solid #ccd8df}}table{{border-collapse:collapse;width:100%}}th,td{{border:1px solid #ccd8df;padding:8px;vertical-align:top}}th{{background:#1f4e78;color:#fff}}</style><h1>{html.escape(title)}：经营者集中申报审查</h1><p>仅基于用户材料的内部初筛；缺失数据不等于未达到门槛，不构成申报义务最终意见。</p><div class="cards"><div class="card">门槛：{html.escape(s["threshold_assessment"])}</div><div class="card">全球营业额：{s["global_total"]} 亿元</div><div class="card">中国境内：{s["china_total"]} 亿元</div><div class="card">抢跑：{html.escape(s["gun_jumping"])}</div></div><h2>风险与建议</h2><table><tr><th>主题</th><th>等级</th><th>发现</th><th>建议</th></tr>{rows}</table><h2>来源与边界</h2><p>交易文件 SHA-256：{b["transaction_sha256"]}。法律来源、控制权和营业额均须按报告待核验项复核。</p>',encoding='utf-8')
 if a.cleanup: cleanup(ws)
 print(json.dumps({'xlsx':str(xlsx),'html':str(htm),'summary':s,'risk_counts':counts},ensure_ascii=False,default=dict))
def main():
 p=argparse.ArgumentParser(); sub=p.add_subparsers(dest='cmd',required=True); q=sub.add_parser('prepare'); q.add_argument('--transaction',required=True); q.add_argument('--attachment',action='append'); q.add_argument('--workspace',required=True); q.add_argument('--processing-environment',choices=['local','cloud'],required=True); q.add_argument('--privacy-confirmed',action='store_true'); q.set_defaults(fn=prepare); q=sub.add_parser('report'); q.add_argument('--bundle',required=True); q.add_argument('--analysis-json',required=True); q.add_argument('--output-dir',required=True); q.add_argument('--cleanup',action='store_true'); q.set_defaults(fn=report); a=p.parse_args()
 try: return a.fn(a) or 0
 except (PipelineError,OSError,ValueError,json.JSONDecodeError,zipfile.BadZipFile) as e: print(f'错误：{e}',file=sys.stderr); return 2
if __name__=='__main__': raise SystemExit(main())
